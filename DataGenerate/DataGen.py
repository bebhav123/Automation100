import openpyxl


def dataGenerator():
   wk = openpyxl.load_workbook("./Excel/tdata.xlsx")
   sh = wk['Sheet1']
   rows = sh.max_row
   li=[]
   lii=[]
   for i in range(1,rows+1):
    lii=[]
    un = sh.cell(i, 1)
    up = sh.cell(i, 2)
    lii.insert(0, un.value)
    lii.insert(1, up.value)
    li.insert(i-1, lii)

   print(li)
   return li